from openpyxl import load_workbook
from abc import ABC, abstractmethod
import numpy as np
from matplotlib import pyplot as plt
from tkinter import filedialog
from tkinter import Tk
from tkinter import PhotoImage
import base64
import os


class RealFunction(ABC):
    @abstractmethod
    def calculate_y(self, x: float):
        pass


class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y


class CubicFunction(RealFunction):
    def __init__(self, a, b, c, d):
        self.a = a
        self.b = b
        self.c = c
        self.d = d

    def calculate_y(self, x: float):
        resultado = (self.a * x ** 3) + (self.b * x ** 2) + (self.c * x) + self.d
        return resultado


def gauss_jordan(m, eps=1.0 / (10 ** 10)):
    """Puts given matrix (2D array) into the Reduced Row Echelon Form.
     Returns True if successful, False if 'm' is singular.
     NOTE: make sure all the matrix items support fractions! Int matrix will NOT work!
     Written by Jarno Elonen in April 2005, released into Public Domain"""
    (h, w) = (len(m), len(m[0]))
    for y in range(0, h):
        maxrow = y
        for y2 in range(y + 1, h):  # Find max pivot
            if abs(m[y2][y]) > abs(m[maxrow][y]):
                maxrow = y2
        (m[y], m[maxrow]) = (m[maxrow], m[y])
        if abs(m[y][y]) <= eps:  # Singular?
            return False
        for y2 in range(y + 1, h):  # Eliminate column y
            c = m[y2][y] / m[y][y]
            for x in range(y, w):
                m[y2][x] -= m[y][x] * c
    for y in range(h - 1, 0 - 1, -1):  # Backsubstitute
        c = m[y][y]
        for y2 in range(0, y):
            for x in range(w - 1, y - 1, -1):
                m[y2][x] -= m[y][x] * m[y2][y] / c
        m[y][y] /= c
        for x in range(h, w):  # Normalize row y
            m[y][x] /= c
    return True


def create_cubic_spline(data_points: list[Point]) -> list[CubicFunction]:
    """
    Data points son los puntos a trazar en forma de lista de tupla.Usar int o float de python.
    Retorna una lista de listas con los coeficientes de cada polinomio, en orden de izquierda a derecha.
    Ej.[[a1, b1, c1, d1], ..., [ai, bi, ci, di]]
    """
    n_puntos = len(data_points)
    n_poly = n_puntos - 1

    # Construir la matriz inicial llena de ceros.
    matriz = []
    for row in range(0, n_poly * 4):
        cero_row = []
        for column in range(0, n_poly * 4 + 1):
            cero_row.append(0)
        matriz.append(cero_row)

    # Por cada punto se generan las ecuaciones restrictivas de las curvas
    for punto in range(0, n_puntos):
        xi = data_points[punto].x
        yi = data_points[punto].y
        fila_f = 0 + 4 * punto  # Fila restricción para que la ecuación a su derecha izquierda cumpla f(x) = y
        fila_f2 = 1 + 4 * punto  # Fila restricción para que la ecuación a su derecha derecha cumpla f(x) = y
        fila_ff = 2 + 4 * punto  # Fila de las ecuaciones fi'(xi) = fi+1'(xi)
        fila_fff = 3 + 4 * punto  # Fila de las ecuaciones fi''(xi) = fi+1''(xi)

        if punto == 0:  # Condición especial de frontera para curva del PRIMER punto.
            # Condición f(x) = y para los puntos extremos.
            matriz[0][0 + punto * 4] = xi ** 3
            matriz[0][1 + punto * 4] = xi ** 2
            matriz[0][2 + punto * 4] = xi
            matriz[0][3 + punto * 4] = 1
            matriz[0][-1] = yi
            # Condición f´´(x) para los puntos extremos.
            matriz[1][0 + punto * 4] = 6 * xi
            matriz[1][1 + punto * 4] = 2
            matriz[1][2 + punto * 4] = 0
            matriz[1][3 + punto * 4] = 0
            matriz[1][-1] = 0  # Condición de frontera.
        elif punto == n_puntos - 1:  # Condición especial de frontera para curva del ÚLTIMO punto.
            # Condición f(x) = y para los puntos extremos.
            matriz[2][0 + (punto - 1) * 4] = xi ** 3
            matriz[2][1 + (punto - 1) * 4] = xi ** 2
            matriz[2][2 + (punto - 1) * 4] = xi
            matriz[2][3 + (punto - 1) * 4] = 1
            matriz[2][-1] = yi
            # Condición f´´(x) para los puntos extremos.
            matriz[3][0 + (punto - 1) * 4] = 6 * xi
            matriz[3][1 + (punto - 1) * 4] = 2
            matriz[3][2 + (punto - 1) * 4] = 0
            matriz[3][3 + (punto - 1) * 4] = 0
            matriz[3][-1] = 0  # Condición de frontera.
        else:  # Condiciones para las curvas que pasan por puntos INTERMEDIOS.
            # La cúbica a la izquierda cumple f(x) = y.
            matriz[fila_f][0 + (punto - 1) * 4] = xi ** 3
            matriz[fila_f][1 + (punto - 1) * 4] = xi ** 2
            matriz[fila_f][2 + (punto - 1) * 4] = xi
            matriz[fila_f][3 + (punto - 1) * 4] = 1
            matriz[fila_f][-1] = yi
            # La cúbica a la derecha cumple f(x) = y
            matriz[fila_f2][0 + (punto) * 4] = xi ** 3
            matriz[fila_f2][1 + (punto) * 4] = xi ** 2
            matriz[fila_f2][2 + (punto) * 4] = xi
            matriz[fila_f2][3 + (punto) * 4] = 1
            matriz[fila_f2][-1] = yi
            # Sus derivadas valen igual
            matriz[fila_ff][0 + (punto - 1) * 4] = 3 * xi ** 2
            matriz[fila_ff][1 + (punto - 1) * 4] = 2 * xi
            matriz[fila_ff][2 + (punto - 1) * 4] = 1
            matriz[fila_ff][0 + (punto) * 4] = -3 * xi ** 2
            matriz[fila_ff][1 + (punto) * 4] = -2 * xi
            matriz[fila_ff][2 + (punto) * 4] = -1
            matriz[fila_ff][-1] = 0
            # Sus segundas derivadas valen igual
            matriz[fila_fff][0 + (punto - 1) * 4] = 6 * xi
            matriz[fila_fff][1 + (punto - 1) * 4] = 2
            matriz[fila_fff][0 + (punto) * 4] = -6 * xi
            matriz[fila_fff][1 + (punto) * 4] = -2
            matriz[fila_fff][-1] = 0

    # Aplicarle la reducción de Gauss-Jordan a la matriz del sistema de ecuaciones.
    gauss_jordan(matriz)

    # Extraer los coeficientes en una lista de listas de coeficientes.
    coeficientes: list = []  # Coeficientes de las ecuaciones en forma [a, b ,c ,d]
    for poly in range(0, n_poly):
        a = matriz[0 + poly * 4][-1]
        b = matriz[1 + poly * 4][-1]
        c = matriz[2 + poly * 4][-1]
        d = matriz[3 + poly * 4][-1]
        coeficientes.append(CubicFunction(a, b, c, d))
    return coeficientes


class CubicSpline(RealFunction):
    def __init__(self, data_points: list[Point]):
        """
        Usar esto para definir Splines Cúbicos.
        El Dominio del Spline es el conjunto cerrado que contiene todos los puntos.
        :param data_points: Lista de objetos Point
        """
        self.functions: list[CubicFunction] = create_cubic_spline(data_points)
        self.interval_boundaries: list[float] = [boundary.x for boundary in data_points] + [data_points[-1].x]
        self.domain: [float, float] = [self.interval_boundaries[0], self.interval_boundaries[-1]]

        mid_l = int(np.floor(len(data_points) / 2))
        mid_r = mid_l + 1
        x = (self.interval_boundaries[mid_l] + self.interval_boundaries[mid_r]) / 2
        self.apex = Point(x, self.calculate_y(x))

    def calculate_y(self, x: float):
        for i, function in enumerate(self.functions):
            """Levanta un error si x está fuera del dominio."""
            left_boundary = self.interval_boundaries[i]
            right_boundary = self.interval_boundaries[i + 1]
            x_within_interval = left_boundary <= x <= right_boundary
            if x_within_interval:
                return function.calculate_y(x)
        raise ValueError("x fuera del dominio. x={}, dominio=[{}, {}]".format(x, self.domain[0], self.domain[1]))


class Constants:
    """Clase que almacena la longitud de del lado de un diagrama ternario y la pendiente de sus lados a y b.
    ¡Nunca modificar nada de acá!."""
    SIDE: float = 1 / (np.sqrt(3) / 2)
    SLOPE_B: float = 1 / (SIDE / 2)
    SLOPE_A: float = -1 / (SIDE / 2)


constants = Constants()


class Terna:
    def __init__(self, a, b, c):
        self.a = a
        self.b = b
        self.c = c

        count_none_components = 0
        for component in [self.a, self.b, self.c]:
            if component is None:
                count_none_components += 1

        if count_none_components > 1:
            pass
            # logging.warning("No se puede crear una Terna con {} coordenadas nulas. Se retornó 0".format(count_none_components))
        elif count_none_components == 1:
            def autofill(comp_a: float | None, comp_b: float | None, comp_c: float | None) -> tuple[float, float, float]:
                if comp_a is None:
                    comp_a = 1 - comp_b - comp_c
                if comp_b is None:
                    comp_b = 1 - comp_a - comp_c
                if comp_c is None:
                    comp_c = 1 - comp_a - comp_b
                return {"a": comp_a,
                        "b": comp_b,
                        "c": comp_c}
            filled = autofill(a, b, c)
            self.a = filled["a"]
            self.b = filled["b"]
            self.c = filled["c"]

        # Normalizar  ->  a + b + c = 1
        total = self.a + self.b + self.c
        self.a = self.a / total
        self.b = self.b / total
        self.c = self.c / total

        ak = constants.SLOPE_B
        bk = -1
        ck = 0
        self.y = self.c
        self.x = ((self.b * np.sqrt(ak * ak + bk * bk)) - bk * self.y - ck) / ak

    def get_xy(self) -> Point:
        """
        Devuelve las coordenadas cartesianas de la terna.
        :return:
        """
        ak = constants.SLOPE_B
        bk = -1
        ck = 0
        y = self.c
        x = ((self.b * np.sqrt(ak * ak + bk * bk)) - bk * y - ck) / ak

        return Point(x, y)

    def __str__(self):
        return "a:{} b:{} c:{}".format(self.a, self.b, self.c)


def main(file_name):
    # file = "datos_binodal.xlsx"
    wb = load_workbook(file_name, read_only=True)
    ws = wb.active
    data = ws['C3':'H102']
    wb.close()

    l = []
    r = []
    row = 0
    valid_line = True
    # Escanear excel
    while valid_line:
        invalid_line = None in [cell.value for cell in data[row][0:6]]
        if invalid_line:
            # print("Invalid line in {}° row. Process finished".format(row + 1))
            break

        A1 = data[row][0].value
        B1 = data[row][1].value
        C1 = data[row][2].value
        a = Terna(A1, B1, C1).get_xy()
        A2 = data[row][3].value
        B2 = data[row][4].value
        C2 = data[row][5].value
        b = Terna(A2, B2, C2).get_xy()
        l.append(a)
        r.append(b)
        row += 1

    # Fig, Ax
    fig = plt.figure(frameon=False, dpi=100)
    w = 9
    h = 9
    fig.set_size_inches(w, h)
    ax = plt.Axes(fig, [0., 0., 1., 1.])
    fig.add_axes(ax)
    margen = 0.05
    plt.xlim(0 - margen, 1.15 + margen)
    plt.ylim(0 - margen - 0.09, 1 + margen)
    plt.gca().set_aspect('equal', adjustable='box')
    ax.set_axis_off()

    r.reverse()
    data_points = l + r

    curva = CubicSpline(data_points)
    x = np.linspace(curva.domain[0], curva.domain[1], 100)
    y = np.array([curva.calculate_y(n) for n in x])
    r.reverse()

    lado = 2 * 1/np.tan(2*np.pi/6)
    mb = 2/lado
    ma = -2/lado

    subdivisiones = 10
    n = subdivisiones + 1
    div_c_x = np.linspace(0, lado, n)
    div_b_x = np.linspace(0, lado/2, n)
    div_a_x = np.linspace(lado/2, lado, n)

    div_c_y = div_c_x * 0
    div_b_y = div_b_x * mb
    div_a_y = div_a_x * ma + 2

    # Diagonales
    for i in range(0, len(div_a_x)):
        k = len(div_a_x)-i-1
        color = "#C0C0C0"
        lw = 0.55
        ax.plot([div_b_x[i], div_c_x[i]], [div_b_y[i], div_c_y[i]], color=color, linewidth=lw)
        ax.plot([div_a_x[i], div_c_x[i]], [div_a_y[i], div_c_y[i]], color=color, linewidth=lw)
        ax.plot([div_a_x[i], div_b_x[k]], [div_a_y[i], div_b_y[k]], color=color, linewidth=lw)

    v_30 = [np.cos(30/360*2*np.pi), np.sin(30/360*2*np.pi)]
    v_150 = [np.cos(150 / 360 * 2 * np.pi), np.sin(150 / 360 * 2 * np.pi)]
    v_270 = [np.cos(3/4 * 2 * np.pi), np.sin(3 / 4 * 2 * np.pi)]

    v_180 = [np.cos(180 / 360 * 2 * np.pi), np.sin(180 / 360 * 2 * np.pi)]
    v_240 = [np.cos(240 / 360 * 2 * np.pi), np.sin(240 / 360 * 2 * np.pi)]
    v_120 = [np.cos(120 / 360 * 2 * np.pi), np.sin(120 / 360 * 2 * np.pi)]

    margin = 0.04
    shift = 0.024
    SA = -0.028
    marc = 0.028
    mara = -0.003
    p0 = [lado/2, 1/3]
    pA = [v_30[0] * 1/3 + p0[0] + v_30[0] * (margin + mara) + v_120[0] * (shift + SA), v_30[1] * 1/3 + p0[1] + v_30[1] * (margin + mara) + v_120[1] * (shift + SA)]
    pB = [v_150[0] * 1/3 + p0[0] + v_150[0] * margin + v_240[0] * shift, v_150[1] * 1/3 + p0[1] + v_150[1] * margin + v_240[1] * shift]
    pC = [v_270[0] * 1/3 + p0[0] + v_270[0] * (margin + marc) + v_180[0] * shift, v_270[1] * 1/3 + p0[1] + v_270[1] * (margin + marc) + v_180[1] * shift]

    # Binodal
    color = "#1C8854"
    lw = 1.8
    ax.plot(x, y, color=color, linewidth=lw)

    # Reparto
    for i in range(0, len(l)):
        # Rectas de reparto
        color = "#1C8854"
        lw = 1.2
        ax.plot([l[i].x, r[i].x], [l[i].y, r[i].y], color=color, linewidth=lw)

        # Puntos de reparto
        color = "#1C8854"
        size = 5
        ax.plot(l[i].x, l[i].y, marker="o", color=color, markersize=size)
        ax.plot(r[i].x, r[i].y, marker="o", color=color, markersize=size)

    # Triangulo
    color = "#252525"
    lw = 2.1
    plt.plot([0, lado], [0, 0], color=color, linewidth=lw)
    plt.plot([0, lado / 2], [0, 1], color=color, linewidth=lw)
    plt.plot([lado / 2, lado], [1, 0], color=color, linewidth=lw)

    # Tags
    color = "#6557d2"
    ax.text(pB[0], pB[1], "B", fontsize=23, color=color, rotation=60, rotation_mode='anchor')
    ax.text(pA[0], pA[1], "A", fontsize=23, color=color, rotation=300, rotation_mode='anchor')
    ax.text(pC[0], pC[1], "C", fontsize=23, color=color, rotation=0, rotation_mode='anchor')

    # Spam
    color = "#9892C9"
    ax.text(0.02, 0.9, "Cortesía de Agustín Tomás Paredes \nig: agustin_t_paredes", color=color)

    return fig


def open_file():
    #icon = "AAABAAEAEBAAAAEAIABoBAAAFgAAACgAAAAQAAAAIAAAAAEAIAAAAAAAAAQAANcNAADXDQAAAAAAAAAAAABe2FAAXthQBV7YUHxe2FDwXthQ/1/YUP9f2FH/XNdN/07TPf9M0jr/TNI6/0zSOv9M0jrxTNI6fU3SOwVN0jsAXthQAF7YUDde2FDsXthQ/17YUP9X10r/VNZH/1vXTf9KzTj/Sc02/0zSOv9M0jr/TNI6/0zSOuxM0jo4TNI6AF7YUABe2FBTXthQ/F/YUf9U1kf/Jc8d/x7NF/9H1Dv/QcAt/0HDLf9N0zv/TNI6/0zSOv9M0jr8TNI6VEzSOgBe2FAAXthQJ17YUN9f2FH/S9U//xfMEf8TzA3/OtIv/07RPP9DxjD/Rsoz/0zSOv9M0jr/TNI64EzSOidM0joAXthQAF7YUABe2FCIXthQ/1vXTf880jH/NdEs/1PWRv9Cwi7/KaMR/yynFf9GyTL/TNM6/0zSOohV1UQATdI7AAAAAABe2FAAXthQKl7YUOBe2E//XdhP/17YUP9d2E7/Pr0p/yafDf8pohD/Q8Yw/03TO+BM0joqTNI6AAAAAAAAAAAAXthQAAD/AACIzYWIlMqU/4DPe/9k11f/W9hM/0vOOf87uyb/Pr8q/0vROf9M0jqJVNVDAE3SOwAAAAAAAAAAAAAAAADAvssAwb7MKr++yeC8v8X/ocak/3DSZ/9P0z3/S9M4/0vTOf9N0jvhTdI7KkzSOgAAAAAAAAAAAAAAAAAAAAAAvb/HAL2/xwC9v8eIvb/H/76/yf+ywLn/ib6G/3TDbf94wnH/e8F1iQD/AABS0EIAAAAAAAAAAAAAAAAAAAAAAAAAAAC9v8cAvb/HKr2/x+C9v8f/vL3F/7CwuP+tr7X/rq+14bKuuiuwrrcAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAvb/HAL2/xwC9v8eLvb/H/7u9xf+vsbb/ra+0/62vtIyztboArrC1AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC9v8cAvb/HXb2/x/67vcX/r7G2/62vtP6tr7Rera+0AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC9v8cAvb/HAL2/x2C9v8f+u73F/6+xtv+tr7T+ra+0YK2vtACusLYAAAAAAAAAAAAAAAAAAAAAAAAAAAC9v8cAvb/HAL2/x0a9v8fOvb/H/7u9xf+vsbb/ra+0/62vtM6tr7RHra+0AK6wtQAAAAAAAAAAAAAAAAAAAAAAvb/HAL2/xwK9v8envb/H/72/x/+7vcX/r7G2/62vtP+tr7T/ra+0qa6wtgOusLUAAAAAAAAAAAAAAAAAAAAAAL2/xwC9v8cAvb/Hbr2/x/e9v8f/u73F/6+xtv+tr7T/ra+0962vtG+srrMArrC1AAAAAAAAAAAAgAEAAIABAACAAQAAgAEAAMADAADAAwAA4AcAAOAHAADwDwAA8A8AAPgfAAD4HwAA+B8AAPAPAADgBwAA8A8AAA=="
    #icondata = base64.b64decode(icon)
    #tempFile = "icon.ico"
    #iconfile = open(tempFile, "wb")
    #iconfile.write(icondata)
    #iconfile.close()
    root = Tk()
    root.withdraw()
    root.iconbitmap("icon.ico")
    #root.wm_iconbitmap(tempFile)
    ## Delete the tempfile
    #os.remove(tempFile)
    filepath = filedialog.askopenfilename(title="Seleccionar plantilla de datos",
                                          filetypes=(("Documentos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if filepath in [None, ""]:
        return None
    return filepath


def save_file(img):
    save_as = filedialog.asksaveasfile(title="Exportar diagrama ternario",
                                       initialfile="Grafico ternario.png",
                                       defaultextension=".png",
                                       filetypes=[("PNG", "*.png")])
    if save_as in [None, ""]:
        return None

    img.savefig(save_as.name, dpi=400)


if __name__ == "__main__":
    TestingMode = False  # Falso: modo aplicación. True: modo testeo rápido.
    if TestingMode:
        file = "datos_binodal.xlsx"
        main(file)
        plt.show()
    elif not TestingMode:
        file = open_file()
        if file not in [None, ""]:
            fig = main(file)
            save_file(fig)
